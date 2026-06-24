"""
Analyze document type fields across CUNY Academic Works OAI-PMH records.

Scans both data/oai_dc/ and data/qdc/ to build a complete picture of where
document type information is stored, what values appear, and where there are gaps.

Output:
  - Terminal summary
  - data/type_field_analysis.xlsx (detailed breakdown)
"""

import glob
import os
import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

OAI_NS = "http://www.openarchives.org/OAI/2.0/"
DC_NS = "http://purl.org/dc/elements/1.1/"

# Type-related fields to check
# In the XML, qdc fields like "thesis.degree.level" appear as <dc:thesis.degree.level>
TYPE_FIELDS_QDC = [
    "type",
    "thesis.degree.publication.type",
    "thesis.degree.level",
    "thesis.degree.name",
]
TYPE_FIELDS_OAI_DC = [
    "type",
]

# Canonical Academic Works document types
CANONICAL_TYPES = {
    "article", "book", "dissertation", "image", "newspaper", "other",
    "report", "review", "activity", "assessment", "assignment",
    "bibliography", "blog", "bookreview", "booksection", "capstone",
    "casestudy", "composition", "conference", "data",
    "doctoral_capstone_project", "doctoral_dissertation", "findingaid",
    "lecture", "lesson", "masters_capstone_project", "masters_thesis",
    "memorandum", "minutes", "newsletter", "paper", "performance",
    "poem", "poster", "presentation", "reference", "response",
    "simulation", "syllabus", "textbook", "thesis", "tutorial",
}

# Excel styling
HEADER_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Arial", size=12)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_filename(filepath: str) -> tuple[str, str]:
    """Extract institution code and set type from a filename.

    Examples:
        publication_gc_etds.xml  -> ('gc', 'etds')
        publication_bb_pubs.xml  -> ('bb', 'pubs')
        publication_pubs.xml     -> ('cuny', 'pubs')   # cross-institution
        publication_gc.xml       -> ('gc', 'general')
    """
    basename = os.path.splitext(os.path.basename(filepath))[0]
    # Remove the "publication_" prefix
    name = basename.replace("publication_", "", 1)

    # Known set-type suffixes (order matters — check longer suffixes first)
    set_types = [
        "_etds_all", "_etds", "_pubs", "_oers", "_arch", "_conf",
        "_studentpubs",
    ]
    set_type = "general"
    for suffix in set_types:
        if name.endswith(suffix):
            set_type = suffix.lstrip("_")
            name = name[: -len(suffix)]
            break

    institution = name if name else "cuny"
    return institution, set_type


def get_dc_field_values(record_elem, field_name: str) -> list[str]:
    """Get all values of a DC field from a record element."""
    values = []
    for elem in record_elem.iter(f"{{{DC_NS}}}{field_name}"):
        text = (elem.text or "").strip()
        if text:
            values.append(text)
    return values


def scan_file(filepath: str, type_fields: list[str]) -> dict:
    """Scan a single XML file and return type field analysis.

    Returns dict with:
        - record_count: total records in the file
        - field_presence: {field_name: count_of_records_with_this_field}
        - field_values: {field_name: Counter of values}
        - gap_identifiers: list of OAI identifiers with no type fields at all
    """
    result = {
        "record_count": 0,
        "field_presence": {f: 0 for f in type_fields},
        "field_values": {f: Counter() for f in type_fields},
        "gap_identifiers": [],
    }

    try:
        tree = ET.parse(filepath)
    except ET.ParseError as e:
        print(f"  WARNING: Could not parse {filepath}: {e}")
        return result

    root = tree.getroot()

    for record in root.iter(f"{{{OAI_NS}}}record"):
        result["record_count"] += 1
        has_any_type = False

        # Get OAI identifier for gap reporting
        identifier_elem = record.find(f".//{{{OAI_NS}}}identifier")
        identifier = identifier_elem.text if identifier_elem is not None else "unknown"

        for field in type_fields:
            values = get_dc_field_values(record, field)
            if values:
                has_any_type = True
                result["field_presence"][field] += 1
                for v in values:
                    result["field_values"][field][v] += 1

        if not has_any_type:
            result["gap_identifiers"].append(identifier)

    return result


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------

def main():
    qdc_dir = os.path.join(DATA_DIR, "qdc")
    oai_dc_dir = os.path.join(DATA_DIR, "oai_dc")

    # Collect results keyed by filename (without directory)
    # Structure: {basename: {format: scan_result, institution, set_type}}
    all_results = {}

    # --- Scan qdc files ---
    print("Scanning qdc files...")
    qdc_files = sorted(glob.glob(os.path.join(qdc_dir, "*.xml")))
    for filepath in qdc_files:
        basename = os.path.basename(filepath)
        institution, set_type = parse_filename(filepath)
        result = scan_file(filepath, TYPE_FIELDS_QDC)
        all_results.setdefault(basename, {
            "institution": institution,
            "set_type": set_type,
        })
        all_results[basename]["qdc"] = result
    print(f"  Scanned {len(qdc_files)} qdc files")

    # --- Scan oai_dc files ---
    print("Scanning oai_dc files...")
    oai_dc_files = sorted(glob.glob(os.path.join(oai_dc_dir, "*.xml")))
    for filepath in oai_dc_files:
        basename = os.path.basename(filepath)
        institution, set_type = parse_filename(filepath)
        result = scan_file(filepath, TYPE_FIELDS_OAI_DC)
        all_results.setdefault(basename, {
            "institution": institution,
            "set_type": set_type,
        })
        all_results[basename]["oai_dc"] = result
    print(f"  Scanned {len(oai_dc_files)} oai_dc files")

    # --- Terminal summary ---
    print()
    print("=" * 70)
    print("FIELD PRESENCE SUMMARY (by institution + set type)")
    print("=" * 70)

    # Group by institution
    by_institution = defaultdict(list)
    for basename, info in sorted(all_results.items()):
        by_institution[info["institution"]].append((basename, info))

    for inst in sorted(by_institution.keys()):
        print(f"\n  {inst.upper()}")
        for basename, info in by_institution[inst]:
            qdc = info.get("qdc", {})
            oai = info.get("oai_dc", {})
            qdc_count = qdc.get("record_count", 0)
            oai_count = oai.get("record_count", 0)
            count = max(qdc_count, oai_count)

            # Build field summary
            fields_present = []
            qdc_presence = qdc.get("field_presence", {})
            oai_presence = oai.get("field_presence", {})

            if oai_presence.get("type", 0) > 0:
                fields_present.append(f"oai_dc:type({oai_presence['type']})")
            if qdc_presence.get("type", 0) > 0:
                fields_present.append(f"qdc:type({qdc_presence['type']})")
            if qdc_presence.get("thesis.degree.publication.type", 0) > 0:
                fields_present.append(
                    f"qdc:thesis.degree.publication.type({qdc_presence['thesis.degree.publication.type']})"
                )
            if qdc_presence.get("thesis.degree.level", 0) > 0:
                fields_present.append(
                    f"qdc:thesis.degree.level({qdc_presence['thesis.degree.level']})"
                )

            field_str = ", ".join(fields_present) if fields_present else "NO TYPE FIELDS"
            print(f"    {basename} ({count} records): {field_str}")

    # --- Value frequency summary ---
    print()
    print("=" * 70)
    print("VALUE FREQUENCIES")
    print("=" * 70)

    # Aggregate across all files
    all_type_values_oai = Counter()
    all_type_values_qdc = Counter()
    all_thesis_pub_type = Counter()
    all_thesis_level = Counter()
    all_thesis_name = Counter()

    for basename, info in all_results.items():
        oai = info.get("oai_dc", {})
        qdc = info.get("qdc", {})
        oai_vals = oai.get("field_values", {})
        qdc_vals = qdc.get("field_values", {})

        all_type_values_oai += oai_vals.get("type", Counter())
        all_type_values_qdc += qdc_vals.get("type", Counter())
        all_thesis_pub_type += qdc_vals.get("thesis.degree.publication.type", Counter())
        all_thesis_level += qdc_vals.get("thesis.degree.level", Counter())
        all_thesis_name += qdc_vals.get("thesis.degree.name", Counter())

    for label, counter in [
        ("oai_dc dc:type", all_type_values_oai),
        ("qdc dc:type", all_type_values_qdc),
        ("qdc thesis.degree.publication.type", all_thesis_pub_type),
        ("qdc thesis.degree.level", all_thesis_level),
        ("qdc thesis.degree.name", all_thesis_name),
    ]:
        if counter:
            print(f"\n  {label}:")
            for value, count in counter.most_common():
                print(f"    {count:>8,}  {value}")

    # --- Reference comparison ---
    print()
    print("=" * 70)
    print("REFERENCE COMPARISON (vs canonical Academic Works types)")
    print("=" * 70)

    # Collect all type values from dc:type and thesis.degree.publication.type
    all_harvested = set()
    for counter in [all_type_values_oai, all_type_values_qdc, all_thesis_pub_type]:
        for value in counter:
            all_harvested.add(value)

    # Normalize for comparison: lowercase, replace spaces with underscores
    def normalize(value: str) -> str:
        return re.sub(r"[\s]+", "_", value.strip().lower())

    matched = set()
    unmatched = []
    for value in sorted(all_harvested):
        norm = normalize(value)
        if norm in CANONICAL_TYPES:
            matched.add(norm)
        else:
            unmatched.append((value, norm))

    unused = sorted(CANONICAL_TYPES - matched)

    print(f"\n  Matched canonical types: {len(matched)}/{len(CANONICAL_TYPES)}")
    if unmatched:
        print(f"\n  Unmapped values ({len(unmatched)}):")
        for orig, norm in unmatched:
            print(f"    \"{orig}\" (normalized: \"{norm}\")")
    if unused:
        print(f"\n  Unused canonical types ({len(unused)}):")
        for t in unused:
            print(f"    {t}")

    # --- Compute gaps once (reused for terminal summary + Excel) ---
    # Records that have no type info in BOTH formats
    gaps_by_file = {}  # {basename: sorted list of identifiers}
    for basename in sorted(all_results.keys()):
        info = all_results[basename]
        qdc = info.get("qdc", {})
        oai = info.get("oai_dc", {})
        qdc_gaps = set(qdc.get("gap_identifiers", []))
        oai_gaps = set(oai.get("gap_identifiers", []))
        both_gaps = qdc_gaps & oai_gaps
        if both_gaps:
            gaps_by_file[basename] = sorted(both_gaps)

    print()
    print("=" * 70)
    print("GAP REPORT (records with no type fields in either format)")
    print("=" * 70)

    if not gaps_by_file:
        print("\n  No gaps found — every record has type info in at least one format.")
    else:
        total_gaps = sum(len(ids) for ids in gaps_by_file.values())
        for basename, identifiers in gaps_by_file.items():
            print(f"\n  {basename}: {len(identifiers)} records with no type in either format")
            for ident in identifiers[:5]:
                print(f"    {ident}")
            if len(identifiers) > 5:
                print(f"    ... and {len(identifiers) - 5} more")
        print(f"\n  Total records with no type in either format: {total_gaps:,}")

    # --- Write Excel ---
    print()
    print("Writing Excel workbook...")
    wb = Workbook()

    def style_header(ws, num_cols: int):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP_ALIGNMENT

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    # Sheet 1: Field Presence
    ws1 = wb.active
    ws1.title = "Field Presence"
    headers = [
        "File", "Institution", "Set Type", "Records",
        "oai_dc: type",
        "qdc: type", "qdc: thesis.degree.publication.type",
        "qdc: thesis.degree.level", "qdc: thesis.degree.name",
    ]
    ws1.append(headers)
    style_header(ws1, len(headers))

    for basename in sorted(all_results.keys()):
        info = all_results[basename]
        qdc = info.get("qdc", {})
        oai = info.get("oai_dc", {})
        count = max(qdc.get("record_count", 0), oai.get("record_count", 0))
        oai_p = oai.get("field_presence", {})
        qdc_p = qdc.get("field_presence", {})

        row = [
            basename,
            info["institution"],
            info["set_type"],
            count,
            oai_p.get("type", 0),
            qdc_p.get("type", 0),
            qdc_p.get("thesis.degree.publication.type", 0),
            qdc_p.get("thesis.degree.level", 0),
            qdc_p.get("thesis.degree.name", 0),
        ]
        ws1.append(row)

    # Apply body font
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = BODY_FONT
    auto_width(ws1)

    # Sheet 2: Value Frequencies
    ws2 = wb.create_sheet("Value Frequencies")
    headers2 = ["Field", "Format", "Value", "Record Count"]
    ws2.append(headers2)
    style_header(ws2, len(headers2))

    for label, fmt, counter in [
        ("dc:type", "oai_dc", all_type_values_oai),
        ("dc:type", "qdc", all_type_values_qdc),
        ("thesis.degree.publication.type", "qdc", all_thesis_pub_type),
        ("thesis.degree.level", "qdc", all_thesis_level),
        ("thesis.degree.name", "qdc", all_thesis_name),
    ]:
        for value, count in counter.most_common():
            ws2.append([label, fmt, value, count])

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(headers2)):
        for cell in row:
            cell.font = BODY_FONT
    auto_width(ws2)

    # Sheet 3: Reference Comparison
    ws3 = wb.create_sheet("Reference Comparison")
    headers3 = ["Category", "Value", "Normalized", "Notes"]
    ws3.append(headers3)
    style_header(ws3, len(headers3))

    for orig, norm in unmatched:
        ws3.append(["Unmapped (in records but not in canonical list)", orig, norm, ""])
    for t in unused:
        ws3.append(["Unused (in canonical list but not in records)", "", t, ""])

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(headers3)):
        for cell in row:
            cell.font = BODY_FONT
    auto_width(ws3)

    # Sheet 4: Gaps
    ws4 = wb.create_sheet("Gaps")
    headers4 = ["File", "Institution", "Set Type", "OAI Identifier"]
    ws4.append(headers4)
    style_header(ws4, len(headers4))

    for basename, identifiers in gaps_by_file.items():
        info = all_results[basename]
        for ident in identifiers:
            ws4.append([basename, info["institution"], info["set_type"], ident])

    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, max_col=len(headers4)):
        for cell in row:
            cell.font = BODY_FONT
    auto_width(ws4)

    # Save
    output_path = os.path.join(DATA_DIR, "type_field_analysis.xlsx")
    wb.save(output_path)
    print(f"Saved to {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
